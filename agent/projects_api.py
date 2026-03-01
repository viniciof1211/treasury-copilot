"""Projects & Contracts BI API — parses Excel data from doc/PROYECTOS."""

import os
import logging
import hashlib
from pathlib import Path
from datetime import datetime, timedelta
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

PROJECTS_DIR = Path(__file__).parent.parent / "doc" / "PROYECTOS"

# ---------------------------------------------------------------------------
# Cache layer — parse Excels once, invalidate on file change
# ---------------------------------------------------------------------------
_cache: dict[str, Any] = {}
_cache_hashes: dict[str, str] = {}


def _file_hash(path: Path) -> str:
    if not path.exists():
        return ""
    return hashlib.md5(path.read_bytes()).hexdigest()


def _cache_valid(key: str, path: Path) -> bool:
    h = _file_hash(path)
    if _cache_hashes.get(key) == h and key in _cache:
        return True
    _cache_hashes[key] = h
    return False


# ---------------------------------------------------------------------------
# ContratosMain.xlsx parser — master contracts
# ---------------------------------------------------------------------------
def _parse_contracts_main() -> list[dict]:
    path = PROJECTS_DIR / "ContratosMain.xlsx"
    cache_key = "contracts_main"
    if _cache_valid(cache_key, path):
        return _cache[cache_key]

    if not path.exists():
        logger.warning(f"ContratosMain.xlsx not found at {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    contracts = []
    for i, r in enumerate(rows):
        if not r or not r[5]:  # skip empty rows (NombreProyecto)
            continue
        try:
            monto_contrato = float(r[11] or 0)
            monto_cancelado = float(r[12] or 0)
            pendiente_cobrar = float(r[13] or 0)
            monto_facturado = float(r[14] or 0)
            pendiente_facturar = float(r[15] or 0)
            adelantos = float(r[16] or 0)

            fecha_inicial = r[2]
            fecha_cierre = r[3]
            fecha_adelanto = r[4]

            contracts.append({
                "id": f"C-{i+1:04d}",
                "eurosat": r[0] or "",
                "consecutivo": r[1],
                "fecha_inicial": fecha_inicial.isoformat() if isinstance(fecha_inicial, datetime) else str(fecha_inicial or ""),
                "fecha_cierre": fecha_cierre.isoformat() if isinstance(fecha_cierre, datetime) else str(fecha_cierre or ""),
                "fecha_adelanto": fecha_adelanto.isoformat() if isinstance(fecha_adelanto, datetime) else str(fecha_adelanto or ""),
                "nombre_proyecto": str(r[5] or "").strip(),
                "proyecto_code": str(r[6] or "").strip(),
                "codigo_cliente": str(r[7] or "").strip(),
                "nombre_cliente": str(r[8] or "").strip(),
                "area": str(r[9] or "").strip(),
                "asesores": str(r[10] or "").strip(),
                "monto_contrato": monto_contrato,
                "monto_cancelado": monto_cancelado,
                "pendiente_cobrar": pendiente_cobrar,
                "monto_facturado": monto_facturado,
                "pendiente_facturar": pendiente_facturar,
                "adelantos": adelantos,
                "empresa": str(r[17] or "").strip(),
                "observaciones": str(r[18] or "").strip(),
                "pct_cobrado": round(monto_cancelado / monto_contrato * 100, 1) if monto_contrato else 0,
                "pct_facturado": round(monto_facturado / monto_contrato * 100, 1) if monto_contrato else 0,
            })
        except Exception as e:
            logger.debug(f"Skipping contract row {i}: {e}")
            continue

    _cache[cache_key] = contracts
    logger.info(f"Parsed {len(contracts)} contracts from ContratosMain.xlsx")
    return contracts


# ---------------------------------------------------------------------------
# tabla cobros.xlsx parser — collections/payments tracking
# ---------------------------------------------------------------------------
def _parse_collections() -> list[dict]:
    path = PROJECTS_DIR / "tabla cobros.xlsx"
    cache_key = "collections"
    if _cache_valid(cache_key, path):
        return _cache[cache_key]

    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["BD"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    collections = []
    for i, r in enumerate(rows):
        if not r or not r[9]:  # skip if no client
            continue
        try:
            fecha_flujo = r[6]
            mes = r[10]
            collections.append({
                "id": f"COL-{i+1:04d}",
                "empresa": str(r[0] or "").strip(),
                "unidad_negocio": str(r[1] or "").strip(),
                "tipo": str(r[2] or "").strip(),
                "recibo_pcgraf": r[3],
                "estado": str(r[4] or "").strip(),
                "semana": str(r[5] or "").strip(),
                "fecha_flujo": fecha_flujo.isoformat() if isinstance(fecha_flujo, datetime) else str(fecha_flujo or ""),
                "moneda": str(r[7] or "").strip(),
                "monto": float(r[8] or 0),
                "cliente": str(r[9] or "").strip(),
                "mes": mes.isoformat() if isinstance(mes, datetime) else str(mes or ""),
                "cobrado": str(r[11] or "").strip(),
                "comentarios": str(r[12] or "").strip(),
            })
        except Exception as e:
            logger.debug(f"Skipping collection row {i}: {e}")
            continue

    _cache[cache_key] = collections
    logger.info(f"Parsed {len(collections)} collections from tabla cobros.xlsx")
    return collections


# ---------------------------------------------------------------------------
# proyecc.xlsx parser — collection projections
# ---------------------------------------------------------------------------
def _parse_projections() -> list[dict]:
    path = PROJECTS_DIR / "proyecc.xlsx"
    cache_key = "projections"
    if _cache_valid(cache_key, path):
        return _cache[cache_key]

    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Detalle"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    projections = []
    for i, r in enumerate(rows):
        if not r or not r[0]:
            continue
        try:
            mes = r[6]
            projections.append({
                "id": f"PRJ-{i+1:04d}",
                "cliente": str(r[0] or "").strip(),
                "monto": float(r[1] or 0),
                "tipo": str(r[2] or "").strip(),
                "area_comercial": str(r[3] or "").strip(),
                "status": str(r[4] or "").strip(),
                "semana": str(r[5] or "").strip(),
                "mes": mes.isoformat() if isinstance(mes, datetime) else str(mes or ""),
            })
        except Exception as e:
            logger.debug(f"Skipping projection row {i}: {e}")
            continue

    _cache[cache_key] = projections
    logger.info(f"Parsed {len(projections)} projections from proyecc.xlsx")
    return projections


# ---------------------------------------------------------------------------
# Analisis cartera parser — aging analysis
# ---------------------------------------------------------------------------
def _parse_aging() -> list[dict]:
    path = PROJECTS_DIR / "Analisis cartera 01julio.xlsx"
    cache_key = "aging"
    if _cache_valid(cache_key, path):
        return _cache[cache_key]

    if not path.exists():
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["General"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    aging = []
    for i, r in enumerate(rows):
        if not r or not r[0]:
            continue
        try:
            fecha_factura = r[12]
            vencimiento = r[13]
            aging.append({
                "id": f"AG-{i+1:04d}",
                "nombre": str(r[0] or "").strip(),
                "tipo_doc": str(r[1] or "").strip(),
                "cod_vendedor": str(r[2] or "").strip(),
                "documento": r[3],
                "negocio": str(r[4] or "").strip(),
                "monto": float(r[5] or 0),
                "sin_vencer": float(r[6] or 0),
                "de_30_dias": float(r[7] or 0),
                "de_60_dias": float(r[8] or 0),
                "de_90_dias": float(r[9] or 0),
                "mas_90_dias": float(r[10] or 0),
                "total": float(r[11] or 0),
                "fecha_factura": fecha_factura.isoformat() if isinstance(fecha_factura, datetime) else str(fecha_factura or ""),
                "vencimiento": vencimiento.isoformat() if isinstance(vencimiento, datetime) else str(vencimiento or ""),
                "dias": int(r[14] or 0),
                "cod_tienda": str(r[15] or "").strip(),
                "empresa": str(r[16] or "").strip(),
                "status": str(r[17] or "").strip(),
                "vendedor": str(r[23] or "").strip() if len(r) > 23 else "",
                "observaciones": str(r[24] or "").strip() if len(r) > 24 else "",
            })
        except Exception as e:
            logger.debug(f"Skipping aging row {i}: {e}")
            continue

    _cache[cache_key] = aging
    logger.info(f"Parsed {len(aging)} aging records from Analisis cartera.xlsx")
    return aging


# ---------------------------------------------------------------------------
# Derived BI calculations
# ---------------------------------------------------------------------------
def _build_projects_portfolio() -> list[dict]:
    """Group contracts by proyecto_code and compute project-level aggregates."""
    contracts = _parse_contracts_main()
    project_map: dict[str, dict] = {}

    for c in contracts:
        key = c["nombre_cliente"]  # group by client
        if key not in project_map:
            project_map[key] = {
                "nombre_cliente": key,
                "contracts": [],
                "total_monto_contrato": 0,
                "total_cancelado": 0,
                "total_pendiente_cobrar": 0,
                "total_facturado": 0,
                "total_pendiente_facturar": 0,
                "total_adelantos": 0,
                "areas": set(),
                "empresas": set(),
                "min_fecha": None,
                "max_fecha": None,
            }
        p = project_map[key]
        p["contracts"].append(c)
        p["total_monto_contrato"] += c["monto_contrato"]
        p["total_cancelado"] += c["monto_cancelado"]
        p["total_pendiente_cobrar"] += c["pendiente_cobrar"]
        p["total_facturado"] += c["monto_facturado"]
        p["total_pendiente_facturar"] += c["pendiente_facturar"]
        p["total_adelantos"] += c["adelantos"]
        if c["area"]:
            p["areas"].add(c["area"])
        if c["empresa"]:
            p["empresas"].add(c["empresa"])

        # Track date range
        for df in ["fecha_inicial", "fecha_cierre"]:
            if c[df]:
                try:
                    dt = datetime.fromisoformat(c[df])
                    if p["min_fecha"] is None or dt < p["min_fecha"]:
                        p["min_fecha"] = dt
                    if p["max_fecha"] is None or dt > p["max_fecha"]:
                        p["max_fecha"] = dt
                except Exception:
                    pass

    # Convert to list with serializable fields
    projects = []
    for key, p in sorted(project_map.items(), key=lambda x: -x[1]["total_monto_contrato"]):
        total = p["total_monto_contrato"]
        projects.append({
            "nombre_cliente": key,
            "contract_count": len(p["contracts"]),
            "total_monto_contrato": round(total, 2),
            "total_cancelado": round(p["total_cancelado"], 2),
            "total_pendiente_cobrar": round(p["total_pendiente_cobrar"], 2),
            "total_facturado": round(p["total_facturado"], 2),
            "total_pendiente_facturar": round(p["total_pendiente_facturar"], 2),
            "total_adelantos": round(p["total_adelantos"], 2),
            "pct_cobrado": round(p["total_cancelado"] / total * 100, 1) if total else 0,
            "pct_facturado": round(p["total_facturado"] / total * 100, 1) if total else 0,
            "areas": sorted(p["areas"]),
            "empresas": sorted(p["empresas"]),
            "fecha_inicio": p["min_fecha"].isoformat() if p["min_fecha"] else None,
            "fecha_fin": p["max_fecha"].isoformat() if p["max_fecha"] else None,
            "contracts": p["contracts"],
        })
    return projects


def _build_milestone_alerts() -> list[dict]:
    """Generate milestone payment alerts for upcoming due dates."""
    contracts = _parse_contracts_main()
    today = datetime.now()
    alerts = []

    for c in contracts:
        if c["pendiente_cobrar"] <= 0 and c["pendiente_facturar"] <= 0:
            continue

        fecha_cierre_str = c.get("fecha_cierre", "")
        if not fecha_cierre_str:
            continue

        try:
            fecha_cierre = datetime.fromisoformat(fecha_cierre_str)
        except Exception:
            continue

        days_until = (fecha_cierre - today).days

        # Only future or slightly past milestones
        if days_until < -30:
            continue

        urgency = "overdue" if days_until < 0 else \
                  "critical" if days_until <= 7 else \
                  "warning" if days_until <= 14 else \
                  "attention" if days_until <= 30 else "ok"

        if urgency == "ok":
            continue

        alerts.append({
            "contract_id": c["id"],
            "nombre_proyecto": c["nombre_proyecto"],
            "nombre_cliente": c["nombre_cliente"],
            "fecha_cierre": fecha_cierre_str,
            "days_until": days_until,
            "urgency": urgency,
            "pendiente_cobrar": c["pendiente_cobrar"],
            "pendiente_facturar": c["pendiente_facturar"],
            "monto_contrato": c["monto_contrato"],
            "area": c["area"],
            "asesores": c["asesores"],
        })

    alerts.sort(key=lambda a: a["days_until"])
    return alerts


def _build_gantt_data(client_filter: str | None = None) -> list[dict]:
    """Build Gantt chart data for contracts with timeline bars."""
    contracts = _parse_contracts_main()
    gantt_items = []

    for c in contracts:
        if client_filter and c["nombre_cliente"].lower() != client_filter.lower():
            continue

        start = c.get("fecha_inicial") or c.get("fecha_adelanto")
        end = c.get("fecha_cierre")
        if not start or not end:
            continue

        try:
            start_dt = datetime.fromisoformat(start)
            end_dt = datetime.fromisoformat(end)
        except Exception:
            continue

        today = datetime.now()
        total_days = max((end_dt - start_dt).days, 1)
        elapsed_days = max((today - start_dt).days, 0)
        progress = min(elapsed_days / total_days * 100, 100)

        # Determine status
        if c["pendiente_cobrar"] <= 0 and c["pendiente_facturar"] <= 0:
            status = "completed"
        elif today > end_dt:
            status = "overdue"
        else:
            status = "active"

        gantt_items.append({
            "id": c["id"],
            "nombre_proyecto": c["nombre_proyecto"],
            "nombre_cliente": c["nombre_cliente"],
            "start": start,
            "end": end,
            "progress": round(progress, 1),
            "status": status,
            "monto_contrato": c["monto_contrato"],
            "pendiente_cobrar": c["pendiente_cobrar"],
            "pendiente_facturar": c["pendiente_facturar"],
            "pct_cobrado": c["pct_cobrado"],
            "area": c["area"],
            "asesores": c["asesores"],
        })

    gantt_items.sort(key=lambda g: g["start"])
    return gantt_items


def _build_area_breakdown() -> list[dict]:
    """Aggregate contract values by business area."""
    contracts = _parse_contracts_main()
    area_map: dict[str, dict] = {}

    for c in contracts:
        area = c["area"] or "Sin Área"
        if area not in area_map:
            area_map[area] = {
                "area": area,
                "count": 0,
                "monto_contrato": 0,
                "cancelado": 0,
                "pendiente_cobrar": 0,
                "facturado": 0,
                "pendiente_facturar": 0,
            }
        a = area_map[area]
        a["count"] += 1
        a["monto_contrato"] += c["monto_contrato"]
        a["cancelado"] += c["monto_cancelado"]
        a["pendiente_cobrar"] += c["pendiente_cobrar"]
        a["facturado"] += c["monto_facturado"]
        a["pendiente_facturar"] += c["pendiente_facturar"]

    result = sorted(area_map.values(), key=lambda x: -x["monto_contrato"])
    for r in result:
        r["pct_cobrado"] = round(r["cancelado"] / r["monto_contrato"] * 100, 1) if r["monto_contrato"] else 0
    return result


def _build_weekly_forecast() -> list[dict]:
    """Build weekly collections forecast from projections."""
    projections = _parse_projections()
    weekly: dict[str, dict] = {}

    for p in projections:
        key = f"{p['mes']}|{p['semana']}"
        if key not in weekly:
            weekly[key] = {
                "mes": p["mes"],
                "semana": p["semana"],
                "total": 0,
                "confirmado": 0,
                "pendiente": 0,
                "count": 0,
            }
        w = weekly[key]
        w["total"] += p["monto"]
        w["count"] += 1
        if p["status"].lower() == "confirmado":
            w["confirmado"] += p["monto"]
        else:
            w["pendiente"] += p["monto"]

    result = sorted(weekly.values(), key=lambda x: (x["mes"], x["semana"]))
    return result


def _build_aging_summary() -> dict:
    """Aggregate aging analysis into buckets."""
    aging = _parse_aging()
    summary = {
        "sin_vencer": 0,
        "de_30_dias": 0,
        "de_60_dias": 0,
        "de_90_dias": 0,
        "mas_90_dias": 0,
        "total": 0,
        "record_count": len(aging),
        "by_negocio": {},
        "by_status": {},
    }

    for a in aging:
        summary["sin_vencer"] += a["sin_vencer"]
        summary["de_30_dias"] += a["de_30_dias"]
        summary["de_60_dias"] += a["de_60_dias"]
        summary["de_90_dias"] += a["de_90_dias"]
        summary["mas_90_dias"] += a["mas_90_dias"]
        summary["total"] += a["total"]

        neg = a["negocio"] or "Sin negocio"
        if neg not in summary["by_negocio"]:
            summary["by_negocio"][neg] = {"total": 0, "count": 0}
        summary["by_negocio"][neg]["total"] += a["monto"]
        summary["by_negocio"][neg]["count"] += 1

        st = a["status"] or "Sin status"
        if st not in summary["by_status"]:
            summary["by_status"][st] = {"total": 0, "count": 0}
        summary["by_status"][st]["total"] += a["monto"]
        summary["by_status"][st]["count"] += 1

    return summary


def _build_kpis() -> dict:
    """Compute top-level KPIs for the projects dashboard."""
    contracts = _parse_contracts_main()
    alerts = _build_milestone_alerts()

    total_contrato = sum(c["monto_contrato"] for c in contracts)
    total_cancelado = sum(c["monto_cancelado"] for c in contracts)
    total_pendiente = sum(c["pendiente_cobrar"] for c in contracts)
    total_facturado = sum(c["monto_facturado"] for c in contracts)
    total_pend_fact = sum(c["pendiente_facturar"] for c in contracts)
    total_adelantos = sum(c["adelantos"] for c in contracts)

    critical_alerts = [a for a in alerts if a["urgency"] in ("critical", "overdue")]
    warning_alerts = [a for a in alerts if a["urgency"] == "warning"]

    unique_clients = len(set(c["nombre_cliente"] for c in contracts))
    unique_areas = len(set(c["area"] for c in contracts if c["area"]))

    return {
        "total_contracts": len(contracts),
        "unique_clients": unique_clients,
        "unique_areas": unique_areas,
        "total_monto_contrato": round(total_contrato, 2),
        "total_cancelado": round(total_cancelado, 2),
        "total_pendiente_cobrar": round(total_pendiente, 2),
        "total_facturado": round(total_facturado, 2),
        "total_pendiente_facturar": round(total_pend_fact, 2),
        "total_adelantos": round(total_adelantos, 2),
        "pct_cobrado_global": round(total_cancelado / total_contrato * 100, 1) if total_contrato else 0,
        "pct_facturado_global": round(total_facturado / total_contrato * 100, 1) if total_contrato else 0,
        "critical_alert_count": len(critical_alerts),
        "warning_alert_count": len(warning_alerts),
        "total_alert_count": len(alerts),
        "critical_alert_value": round(sum(a["pendiente_cobrar"] for a in critical_alerts), 2),
    }


# ---------------------------------------------------------------------------
# Starlette endpoint handlers
# ---------------------------------------------------------------------------
from starlette.requests import Request
from starlette.responses import JSONResponse


async def projects_kpis(request: Request):
    """GET /projects/kpis — top-level project KPIs."""
    try:
        return JSONResponse(_build_kpis())
    except Exception as e:
        logger.error(f"projects_kpis error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_portfolio(request: Request):
    """GET /projects/portfolio — projects grouped by client."""
    try:
        return JSONResponse({"projects": _build_projects_portfolio()})
    except Exception as e:
        logger.error(f"projects_portfolio error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_contracts(request: Request):
    """GET /projects/contracts — all contracts with optional filters."""
    try:
        area = request.query_params.get("area")
        empresa = request.query_params.get("empresa")
        cliente = request.query_params.get("cliente")
        contracts = _parse_contracts_main()
        if area:
            contracts = [c for c in contracts if c["area"].lower() == area.lower()]
        if empresa:
            contracts = [c for c in contracts if c["empresa"].lower() == empresa.lower()]
        if cliente:
            contracts = [c for c in contracts if cliente.lower() in c["nombre_cliente"].lower()]
        return JSONResponse({"contracts": contracts, "total": len(contracts)})
    except Exception as e:
        logger.error(f"projects_contracts error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_alerts(request: Request):
    """GET /projects/alerts — milestone payment alerts."""
    try:
        return JSONResponse({"alerts": _build_milestone_alerts()})
    except Exception as e:
        logger.error(f"projects_alerts error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_gantt(request: Request):
    """GET /projects/gantt — Gantt chart data, optional ?client=X."""
    try:
        client = request.query_params.get("client")
        return JSONResponse({"items": _build_gantt_data(client)})
    except Exception as e:
        logger.error(f"projects_gantt error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_area_breakdown(request: Request):
    """GET /projects/areas — breakdown by business area."""
    try:
        return JSONResponse({"areas": _build_area_breakdown()})
    except Exception as e:
        logger.error(f"projects_area_breakdown error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_collections(request: Request):
    """GET /projects/collections — all collection records."""
    try:
        cliente = request.query_params.get("cliente")
        collections = _parse_collections()
        if cliente:
            collections = [c for c in collections if cliente.lower() in c["cliente"].lower()]
        return JSONResponse({"collections": collections, "total": len(collections)})
    except Exception as e:
        logger.error(f"projects_collections error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_forecast(request: Request):
    """GET /projects/forecast — weekly collections forecast."""
    try:
        return JSONResponse({"forecast": _build_weekly_forecast()})
    except Exception as e:
        logger.error(f"projects_forecast error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_aging(request: Request):
    """GET /projects/aging — aging analysis summary + detail."""
    try:
        summary = _build_aging_summary()
        detail = request.query_params.get("detail", "false").lower() == "true"
        result = {"summary": summary}
        if detail:
            result["records"] = _parse_aging()
        return JSONResponse(result)
    except Exception as e:
        logger.error(f"projects_aging error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)


async def projects_curation_save(request: Request):
    """POST /projects/curation — save curated contract data."""
    try:
        body = await request.json()
        contract_id = body.get("contract_id")
        changes = body.get("changes", {})
        if not contract_id or not changes:
            return JSONResponse({"error": "contract_id and changes required"}, status_code=400)

        # For now, log the curation request — in production this would persist
        logger.info(f"Curation save: contract={contract_id}, changes={changes}")
        return JSONResponse({
            "status": "ok",
            "contract_id": contract_id,
            "changes_applied": changes,
            "note": "Changes logged. Persist to Supabase/ERP in production.",
        })
    except Exception as e:
        logger.error(f"projects_curation_save error: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)
